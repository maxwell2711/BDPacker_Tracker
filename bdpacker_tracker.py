import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side
from datetime import datetime
import random


def process_file(input_path: str, output_path: str) -> None:
    """Read the input Excel file, compute summaries and write them to output_path."""
    # read with pandas; engine openpyxl is default for .xlsx
    df = pd.read_excel(input_path)

    # normalize column names (strip spaces)
    df.columns = df.columns.str.strip()

    # convert date and time columns
    if "date packed" in df.columns:
        df["date packed"] = pd.to_datetime(df["date packed"]).dt.date
    else:
        raise KeyError("'date packed' column not found in input file")

    if "packing time" in df.columns:
        df["packing time"] = pd.to_timedelta(df["packing time"])
    else:
        raise KeyError("'packing time' column not found in input file")

    # ensure est net profit is numeric
    if "est net profit" in df.columns:
        df["est net profit"] = pd.to_numeric(df["est net profit"], errors="coerce").fillna(0)
    else:
        raise KeyError("'est net profit' column not found in input file")

    # clean packer column
    if "packed by" in df.columns:
        df["packed by"] = df["packed by"].astype(str).str.strip()
    else:
        raise KeyError("'packed by' column not found in input file")

    # --- Data anomaly corrections ---
    # Determine a customer identifier: prefer 'email' if present, else combine first+last
    if "email" in df.columns:
        df["_customer_id"] = df["email"].astype(str).str.strip()
    else:
        fn = df.columns.str.strip().str.lower()
        if "first name" in fn and "last name" in fn:
            # map to actual column names
            first_col = df.columns[fn.get_loc("first name")]
            last_col = df.columns[fn.get_loc("last name")]
            df["_customer_id"] = (df[first_col].astype(str).str.strip() + " " +
                                   df[last_col].astype(str).str.strip())
        else:
            # fallback: use index as id (no merging will be attempted)
            df["_customer_id"] = df.index.astype(str)

    # Fix zero packing times by copying a non-zero packing time for same customer/date
    zero_mask = df["packing time"] == pd.Timedelta(0)
    if zero_mask.any():
        # create lookup of (date, customer) -> first non-zero packing time
        lookup = {}
        for _, row in df.loc[~zero_mask].iterrows():
            key = (row["date packed"], row["_customer_id"])
            if key not in lookup and pd.notna(row["packing time"]):
                if row["packing time"] > pd.Timedelta(0):
                    lookup[key] = row["packing time"]

        # apply to zero rows
        for idx, row in df.loc[zero_mask].iterrows():
            key = (row["date packed"], row["_customer_id"])
            if key in lookup:
                df.at[idx, "packing time"] = lookup[key]

    # Correct estimated labor cost for customers with multiple orders on same date
    # User-provided column name for anomalous labor cost (if present)
    labor_col = "est labor cost"
    if labor_col not in df.columns:
        # create column if missing
        df[labor_col] = 0.0
    # ensure numeric
    try:
        df[labor_col] = pd.to_numeric(df[labor_col], errors="coerce").fillna(0.0)
    except Exception:
        df[labor_col] = 0.0

    # identify customer/date groups with multiple orders
    grp_counts = df.groupby(["date packed", "_customer_id"]).size()

    for (date_val, cust), count in grp_counts.items():
        if count > 1:
            mask = (df["date packed"] == date_val) & (df["_customer_id"] == cust)
            # compute corrected labor per row (hours * $25)
            hours = df.loc[mask, "packing time"].dt.total_seconds() / 3600.0
            corrected_labor = hours * 25.0
            # read old anomalous labor (may be zero)
            old_labor = df.loc[mask, labor_col].fillna(0.0)
            # Adjust est net profit: add back old labor then subtract corrected labor
            if "est net profit" in df.columns:
                df.loc[mask, "est net profit"] = (
                    df.loc[mask, "est net profit"].fillna(0.0) + old_labor - corrected_labor
                )
            # store corrected labor
            df.loc[mask, labor_col] = corrected_labor.values


    # group by date and packer
    grp = df.groupby(["date packed", "packed by"], as_index=False)
    summary = grp["packing time"].sum()
    summary = summary.rename(columns={"packing time": "total_packing_time"})

    # add additional columns for hours/minutes
    summary["total_hours"] = summary["total_packing_time"].dt.total_seconds() / 3600
    summary["total_minutes"] = summary["total_packing_time"].dt.total_seconds() / 60

    # second summary: combine profit information
    profit = grp.agg({"packing time": "sum", "est net profit": "sum"})
    profit = profit.rename(columns={"packing time": "total_packing_time", "est net profit": "total_profit"})
    profit["total_hours"] = profit["total_packing_time"].dt.total_seconds() / 3600
    profit["total_minutes"] = profit["total_packing_time"].dt.total_seconds() / 60
    # avoid division by zero
    profit["profit_per_hour"] = profit.apply(lambda row: row["total_profit"] / row["total_hours"]
                                              if row["total_hours"] > 0 else 0,
                                              axis=1)
    profit["profit_per_minute"] = profit.apply(lambda row: row["total_profit"] / row["total_minutes"]
                                                if row["total_minutes"] > 0 else 0,
                                                axis=1)

    # compute overall (not per packer) daily totals
    daily = df.groupby("date packed").agg(
        total_packing_time=("packing time", "sum"),
        total_profit=("est net profit", "sum"),
    )
    daily["total_hours"] = daily["total_packing_time"].dt.total_seconds() / 3600
    daily["total_minutes"] = daily["total_packing_time"].dt.total_seconds() / 60
    daily["profit_per_hour"] = daily.apply(lambda r: r["total_profit"] / r["total_hours"]
                                             if r["total_hours"] > 0 else 0,
                                             axis=1)
    daily["profit_per_minute"] = daily.apply(lambda r: r["total_profit"] / r["total_minutes"]
                                               if r["total_minutes"] > 0 else 0,
                                               axis=1)
    daily = daily.reset_index()
    daily.insert(1, "packed by", "<all>")

    # append overall daily rows to profit summary
    profit = pd.concat([profit, daily], ignore_index=True, sort=False)
    
    # remove unnecessary columns before writing
    summary = summary.drop(columns=["total_packing_time"], errors="ignore")
    profit = profit.drop(columns=["total_packing_time", "total_hours"], errors="ignore")

    # write to excel with two sheets
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Packing Time Summary", index=False)
        profit.to_excel(writer, sheet_name="Profit Rate Summary", index=False)

    # format the workbook with colors and separators
    _format_workbook(output_path, df)


def _hex_to_rgb(hex_color: str) -> tuple:
    """Convert hex color (RRGGBB) to RGB tuple."""
    hex_color = hex_color.lstrip('#')
    return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))


def _rgb_to_hex(r: int, g: int, b: int) -> str:
    """Convert RGB tuple to hex color (RRGGBB)."""
    r = max(0, min(255, int(r)))
    g = max(0, min(255, int(g)))
    b = max(0, min(255, int(b)))
    return f"{r:02X}{g:02X}{b:02X}"


def _apply_color_offset(base_hex: str, offset_r: int, offset_g: int, offset_b: int) -> str:
    """Apply RGB offset to a base hex color."""
    r, g, b = _hex_to_rgb(base_hex)
    return _rgb_to_hex(r + offset_r, g + offset_g, b + offset_b)


def _format_workbook(output_path: str, df: pd.DataFrame) -> None:
    """Apply formatting (colors, date separators, column widths) to the output workbook."""
    wb = load_workbook(output_path)
    
    # Base color palette (cycle through for dates)
    base_colors = ["FFE6F0", "E6F0FF", "F0FFE6", "FFF9E6", "F0E6FF", "FFE6E6", "E6FFFF"]
    
    # Get unique dates in order
    unique_dates = sorted(df["date packed"].unique())
    date_to_base_color = {date: base_colors[i % len(base_colors)] for i, date in enumerate(unique_dates)}
    
    # Assign each packer a lightness offset (HLS) to vary shades of base hue
    unique_packers = sorted(df["packed by"].unique())
    random.seed(42)
    # create small H/S/L offsets so packers remain in the same colour family
    def _random_offsets():
        return (random.uniform(-0.03, 0.03),  # hue offset (small)
                random.uniform(-0.12, 0.12),  # saturation offset
                random.uniform(-0.06, 0.06))  # lightness offset
    packer_offsets = {packer: _random_offsets() for packer in unique_packers}
    
    # Process both sheets
    for sheet_name in ["Packing Time Summary", "Profit Rate Summary"]:
        ws = wb[sheet_name]
        
        # Auto-fit column widths
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
        
        # Build list of (row_idx, date) for easy lookup
        date_rows = []
        for row_idx in range(2, ws.max_row + 1):
            date_val = ws.cell(row=row_idx, column=1).value
            date_rows.append((row_idx, date_val))
        
        # Determine which rows are the last entry for each date
        last_row_of_date = set()
        for i in range(len(date_rows)):
            current_date = date_rows[i][1]
            if i == len(date_rows) - 1 or date_rows[i + 1][1] != current_date:
                last_row_of_date.add(date_rows[i][0])
        
        # Apply minimal borders and colors
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        thick_bottom = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thick")
        )
        
        for row_idx in range(2, ws.max_row + 1):
            date_val = ws.cell(row=row_idx, column=1).value
            packer_val = ws.cell(row=row_idx, column=2).value
            
            # normalize the date we read from the sheet; openpyxl may return a datetime
            if hasattr(date_val, "date"):
                date_val = date_val.date()
            # if it's a string parse it just in case
            if isinstance(date_val, str):
                try:
                    date_val = datetime.fromisoformat(date_val).date()
                except Exception:
                    pass
            # Get base color for this date (fallback to white)
            base_color = date_to_base_color.get(date_val, "FFFFFF")
            
            # Apply HLS offsets per packer so shades stay in the same family
            if packer_val == "<all>":
                color = "FFFFFF"
            else:
                offset_h, offset_s, offset_l = packer_offsets.get(packer_val, (0, 0, 0))
                r, g, b = _hex_to_rgb(base_color)
                import colorsys
                h, l, s = colorsys.rgb_to_hls(r/255.0, g/255.0, b/255.0)
                # prevent extremely-high lightness from washing out variations
                if l > 0.92:
                    l = 0.90
                # apply small offsets and clamp
                h = (h + offset_h) % 1.0
                s = max(0, min(1, s + offset_s))
                l = max(0, min(1, l + offset_l))
                r2, g2, b2 = colorsys.hls_to_rgb(h, l, s)
                color = _rgb_to_hex(int(r2*255), int(g2*255), int(b2*255))
            
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            
            is_last_of_date = row_idx in last_row_of_date
            border = thick_bottom if is_last_of_date else thin_border
            
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = fill
                cell.border = border
    
    wb.save(output_path)


class App:
    def __init__(self, root: tk.Tk):
        self.root = root
        root.title("BDPacker Tracker")
        root.geometry("500x200")

        self.filename = tk.StringVar()
        self.outputname = tk.StringVar()

        tk.Label(root, text="Input Excel file:").pack(pady=(20, 0))
        tk.Entry(root, textvariable=self.filename, width=60).pack(padx=10)
        tk.Button(root, text="Browse...", command=self.browse_file).pack(pady=5)

        tk.Button(root, text="Process", command=self.run).pack(pady=10)
        tk.Label(root, textvariable=self.outputname, fg="green").pack()

    def browse_file(self):
        path = filedialog.askopenfilename(
            title="Select Excel file",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*")],
        )
        if path:
            self.filename.set(path)

    def run(self):
        inp = self.filename.get()
        if not inp:
            messagebox.showwarning("No file", "Please select an input Excel file first.")
            return
        
        # read data to get date range for default filename
        try:
            temp_df = pd.read_excel(inp)
            temp_df.columns = temp_df.columns.str.strip()
            temp_df["date packed"] = pd.to_datetime(temp_df["date packed"]).dt.date
            dates = temp_df["date packed"].unique()
            min_date = min(dates)
            max_date = max(dates)
            
            # format filename based on date range (Windows-compatible)
            def format_date(d):
                s = d.strftime("%m-%d-%y")
                return s.lstrip('0').replace('-0', '-')
            
            if min_date == max_date:
                date_str = format_date(min_date)
                default_filename = f"summary_output_{date_str}.xlsx"
            else:
                min_str = format_date(min_date)
                max_str = format_date(max_date)
                default_filename = f"summary_output_{min_str}_{max_str}.xlsx"
        except Exception:
            default_filename = "summary_output.xlsx"
        
        # ask where to save
        default_out = os.path.join(os.path.dirname(inp), default_filename)
        out = filedialog.asksaveasfilename(
            title="Save summary as",
            defaultextension=".xlsx",
            initialfile=default_filename,
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not out:
            return
        try:
            process_file(inp, out)
            self.outputname.set(f"Created: {out}")
            messagebox.showinfo("Success", "Summary file written successfully.")
        except Exception as e:
            messagebox.showerror("Error", str(e))


if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()
